# -*- coding: utf-8 -*-
"""
build_crosstalk_v03.py —— 生成 crosstalk_sim_v03.ipynb
========================================================
基于 Elephant 时序表.xlsx 的长焦时序（只看 Group A）。
改进 v01：
  · 16 通道回波接收图（色块/横线，更清晰）
  · 时序图用色块条展示（而非离散散点）
  · 每个激光器多次发光的累积 TOF 波形（修正时间计算 + 信号来源标记）
  · 去掉 B 组（只看 A 组）
  · 时间计算：echo_time - fire_time（相对时间），非绝对时间
  · 每个信号标明来自哪个激光器
"""
import json

CELL_HEADER = r'''# 串扰仿真 v02 —— 短焦 16 激光器编码串扰检测（基于 Elephant 时序，只看 Group A）

> 理想回波模型（δ 函数、100% 探测率），不考虑蒙卡与波形展宽。
> 时序图从 `Elephant 时序表.xlsx` 长焦 tab 的 Group A 读取。

## 核心概念
- **Kick（触发节拍）**：Group A 的 16 个 kick，每个 kick 有 16 个激光器发光。
- **编码**：`tx_trig_dly`（1ns 步长）+ `tdelay`（1/12ns 步长）。
- **TOF 窗**：固定 2000ns（0.2us）。
- **串扰鬼影**：激光器 A 的回波落入激光器 B 的 TOF 窗，B 按自己的 `echo_time - B_fire_time` 计算距离。

> 缩写：TOF（Time of Flight，飞行时间）。
'''

CELL_IMPORTS = r'''import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import openpyxl

for _f in ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Source Han Sans SC"]:
    try:
        matplotlib.rcParams["font.sans-serif"] = [_f]; break
    except Exception:
        continue
matplotlib.rcParams["axes.unicode_minus"] = False

C_LIGHT = 2.99792458e8
NS = 1e-9
print("模块导入完成。")
'''

CELL_EXCEL_PARSE = r'''# ============================================================================
# 从 Excel 读取 Group A 时序（只看 A 组）
# ============================================================================
EXCEL_FILE = "Elephant 时序表.xlsx"
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb[wb.sheetnames[0]]  # 第一个 tab = 长焦

# ---- 解析行 1-6 的标头 ----
# 第 4 行: tdelay (1/12ns 步长)
tdelay_A = {}
for c in range(8, 24):  # Group A 列 8-23
    v = ws.cell(4, c).value
    if v is not None: tdelay_A[c - 8] = v

# 第 5 行: gate (ns)
gate_A = {}
for c in range(8, 24):
    v = ws.cell(5, c).value
    if v is not None: gate_A[c - 8] = v

# ---- 解析激光器数据行 (7~85) ----
# 每行: C1=DSP cluster, C2=HDC, C3=TX_TRG, C4=RX_TRG, C5=CHG, C6=LASER
# 然后列 8-23 (组A kick 0~15) 标记是否发光
# 非空 = 发激光; 值(0/50/...)就是 tx_trig_dly (1ns 步长)

laser_data = []
for r in range(7, 86):
    laser_id = ws.cell(r, 6).value
    if laser_id is None: continue
    try: laser_id = int(laser_id)
    except (ValueError, TypeError): continue
    fires = []
    for c in range(8, 24):  # Group A 列
        v = ws.cell(r, c).value
        if v is not None:
            fires.append((c - 8, int(v)))  # (kick_idx, tx_trig_dly)
    laser_data.append({
        "laser_id": laser_id,
        "tx_trg": str(ws.cell(r, 3).value or ""),
        "fires": fires,
    })

N_LASERS = len(laser_data)
print(f"从 {EXCEL_FILE} 读取 Group A: {N_LASERS} 个激光器")
print(f"  {'Laser':>6} {'TX_TRG':>6} {'发光次数':>8} {'Kick(tx_trig_dly)'}")
print(f"  {'-'*6} {'-'*6} {'-'*8} {'-'*30}")
for ld in laser_data:
    ks = ",".join(f"K{k}={d}" for (k, d) in ld["fires"])
    print(f"  {ld['laser_id']:>6d} {ld['tx_trg']:>6} {len(ld['fires']):>8d} {ks}")
print(f"\n总激光器: {N_LASERS}, 总发光事件: {sum(len(ld['fires']) for ld in laser_data)}")
'''

CELL_FIRING_TIMES = r'''# ============================================================================
# 计算发光时刻（只看 Group A）
# ============================================================================
# 每个激光器的总偏移 = tx_trig_dly*1ns + tdelay[kick_idx]/12ns + fpga_jitter*8ns
# KICK_SPACING = 2.2us（两个 tick 间隔），TOF = 2000ns（固定）

KICK_SPACING = 2.2e-6        # 相邻 kick 间隔 [s]（2.2us）
KICK_BASE_START = 0.0         # 第一个 kick 的起始时刻 [s]

# FPGA 抖动暂设 0（可调）
FPGA_JITTER = np.zeros(N_LASERS + 1)   # {laser_id: 8ns步长数}

def laser_fire_offset(kick_idx, tx_trig_dly):
    """该激光器在该 kick 中的总偏移 [s] = tx_trig_dly·1ns + tdelay[kick_idx]/12ns"""
    td = tdelay_A.get(kick_idx, 0) * (1e-9 / 12.0)
    return tx_trig_dly * 1e-9 + td

def kick_base_time(kick_idx):
    return KICK_BASE_START + kick_idx * KICK_SPACING

def kick_tof_us(kick_idx):
    return 2.0  # 固定 TOF = 2000ns（0.2us）

# ---- 计算所有发光事件 ----
firing_events = []  # (laser_id, kick_idx, tx_trig_dly, fire_time[s])
for ld in laser_data:
    lid = ld["laser_id"]
    for (kidx, tr) in ld["fires"]:
        base = kick_base_time(kidx)
        off = laser_fire_offset(kidx, tr)
        t_fire = base + off
        firing_events.append((lid, kidx, tr, t_fire))

firing_events.sort(key=lambda x: x[3])
print(f"共 {len(firing_events)} 次发光事件 (按时间升序):")
print(f"  {'Laser':>6} {'Kick':>5} {'tx_trig_dly':>12} {'FireTime[ns]':>15}")
for lid, kidx, tr, tf in firing_events:
    print(f"  {lid:>6d} {kidx:>5d} {tr:>12d} {tf/NS:>15.3f}")'''

CELL_ECHO_SCAN = r'''# ============================================================================
# 16 通道回波接收：每个激光器收到哪些回波
# ============================================================================
# 对给定目标距离 D，遍历所有发光事件，检查被哪些激光器接收。
# 接收判断：激光器 B 在自己的 TOF 窗 [fire_time, fire_time + 2us] 内收到回波
# TOF 窗 = [B 的发光时刻, B 的发光时刻 + 2us]

def detect_echoes_for_target(D, verbose=False):
    """对距离 D 处的目标，返回所有回波记录。
    每项: ( emitter_laser, kick_idx, fire_time,
            detector_laser, det_kick_idx, det_fire_time,
            echo_time, calc_distance )"""
    t_tof = 2.0 * D / C_LIGHT
    results = []
    for (emit_lid, kidx_e, tr_e, t_fire_e) in firing_events:
        t_echo = t_fire_e + t_tof
        for ld in laser_data:
            det_lid = ld["laser_id"]
            det_fire = 0.0; det_kidx = -1
            for (dl, dk, dtr, dtf) in firing_events:
                if dl == det_lid and dtf <= t_echo:
                    det_fire = dtf; det_kidx = dk
            tof_win_end = det_fire + 2.0e-6  # 固定 TOF = 2us
            if det_fire > 0 and det_fire <= t_echo <= tof_win_end:
                calc_dist = (t_echo - det_fire) * C_LIGHT / 2.0
                results.append((emit_lid, kidx_e, t_fire_e,
                                det_lid, det_kidx, det_fire,
                                t_echo, calc_dist))
    return results

# ---- 测试：一个目标 ----
D_test = 100.0
res = detect_echoes_for_target(D_test)
print(f"测试目标: D={D_test:.0f}m")
print(f"  共 {len(res)} 条回波记录")
emit_set = set(r[0] for r in res)
det_set  = set(r[4] for r in res)
print(f"  发射激光器数: {len(emit_set)}")
print(f"  接收激光器数: {len(det_set)}")
if len(res) > 0:
    print(f"  首条: 发射 Laser{res[0][0]} → 接收 Laser{res[0][3]}, 计算距离={res[0][7]:.2f}m")'''

CELL_16CH_ECHO = r'''# ============================================================================
# 绘图 1：16 通道回波接收图 —— 每个通道收到的信号位置
# ============================================================================
# 横轴 = 时间 [ns]，纵轴 = 16 个激光器通道
# 用色块/横线表示每个通道在不同时刻收到回波
# 颜色 = 发射激光器（区分不同光源）
# 每条回波线从该激光器的发光时刻到回波时刻，计算为 (echo_time - fire_time)

D_SHOW = 150.0       # 示例距离（可调）
res_show = detect_echoes_for_target(D_SHOW)

fig, ax = plt.subplots(figsize=(16, 6))

# 为每个发射激光器分配颜色
emit_ids = sorted(set(r[0] for r in res_show))
emit_colors = {eid: plt.cm.tab20(i / max(1, len(emit_ids) - 1))
               for i, eid in enumerate(emit_ids)}

# 画每个接收通道的"时间线"
for det_lid in range(1, N_LASERS + 1):
    # 该通道收到的回波
    echoes = [r for r in res_show if r[3] == det_lid]
    if not echoes:
        # 空通道画一条浅灰线
        ax.axhline(det_lid - 0.33, det_lid + 0.33, color="0.9", lw=0.5)
        continue
    for (emit_lid, kidx_e, _t_fire_e, _det_lid, _det_kidx, det_fire, t_echo, calc_d) in echoes:
        color = emit_colors.get(emit_lid, "gray")
        # 画水平线段：从该探测器发光时刻到回波时刻（相对时间）
        ax.hlines(det_lid, det_fire, t_echo, colors=color, lw=2.5, alpha=0.8)
        # 在回波时刻画圆点
        ax.plot(t_echo, det_lid, "o", color=color, ms=6, zorder=5)
        # 标明该信号来自哪个激光器（在图例或附近标注）
        if det_lid == 1:  # 只在第一个通道标一次
            ax.text(t_echo, det_lid + 0.5, f"来自 Laser{emit_lid}",
                    fontsize=6, color=color, ha="left", va="bottom")

    # 横轴改为 ns 显示（原坐标是秒，此处转换为 ns 显示）
    import matplotlib.ticker as ticker
    def _sec_to_ns(x, pos): return f"{x*1e9:.0f}"
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(_sec_to_ns))

# 标注
ax.set_xlabel("时间 t [ns]")
ax.set_ylabel("激光器通道 (接收)")
ax.set_yticks(range(1, N_LASERS + 1))
ax.set_ylim(0.5, N_LASERS + 0.5)
ax.set_title(f"16 通道回波接收图：D={D_SHOW:.0f}m"
             f"（颜色=发射激光器，横线=发光→回波时间差，圆点=回波时刻）")
ax.grid(alpha=0.3, axis="x")

# 图例（发射器颜色）
for eid in emit_ids:
    ax.plot([], [], "-", color=emit_colors[eid], lw=2, label=f"Laser {eid} 发射")
ax.legend(fontsize=7, ncol=4, loc="upper right")
plt.tight_layout()
plt.savefig("crosstalk_v03_echo_16ch.png", dpi=110, bbox_inches="tight")
plt.show()

print(f"回波图: D={D_SHOW:.0f}m, {len(res_show)} 条回波, "
      f"{len(emit_ids)} 个发射器")
'''

CELL_TIMING_BLOCK = r'''# ============================================================================
# 绘图 2：时序图 —— 色块条展示（Group A）
# ============================================================================
fig, ax = plt.subplots(figsize=(16, 5.5))

for (lid, kidx, tr, tf) in firing_events:
    tn = tf / NS
    ax.barh(lid, width=8, left=tn, height=0.7, alpha=0.8,
            color=plt.cm.tab20((lid - 1) / 16.0), edgecolor="k", linewidth=0.3)

# 标注 kick 基线
for k in range(16):
    base = kick_base_time(k) / NS
    ax.axvline(base, color="gray", ls=":", lw=0.5, alpha=0.4)
    if k % 2 == 0:
        ax.text(base, N_LASERS + 0.8, f"K{k}", fontsize=6, ha="center", color="gray")
    ax.text(base, N_LASERS + 1.8, f"2us", fontsize=5, ha="center", color="orange")

ax.set_xlabel("时间 t [ns]")
ax.set_ylabel("激光器编号")
ax.set_yticks(range(1, N_LASERS + 1))
ax.set_ylim(0.5, N_LASERS + 3)
ax.set_title("激光器发光时序图（Group A，色块 = 发光事件，灰色虚线 = kick 基线）", fontsize=12)
ax.grid(alpha=0.2, axis="x")
plt.tight_layout()
plt.savefig("crosstalk_v03_timing_blocks.png", dpi=110, bbox_inches="tight")
plt.show()'''

CELL_CUMULATIVE_TOF = r'''# ============================================================================
# 绘图 3：全部 16 激光器 x 1ns 直方图累积 TOF 波形
# ============================================================================
D_CUMUL = 150.0
res_cumul = detect_echoes_for_target(D_CUMUL)

bin_edges = np.arange(0, 2001) * 1e-9
bin_ctrs  = 0.5 * (bin_edges[:-1] + bin_edges[1:])
dist_axis = bin_ctrs * C_LIGHT / 2.0

hist_all = {}
for det_lid in range(1, N_LASERS + 1):
    echoes = [r for r in res_cumul if r[3] == det_lid]
    if not echoes:
        hist_all[det_lid] = (np.zeros(len(bin_ctrs)), np.zeros(len(bin_ctrs)))
        continue
    rel_times = np.array([r[6] - r[5] for r in echoes])
    is_self   = np.array([r[0] == det_lid for r in echoes])
    counts, _ = np.histogram(rel_times, bins=bin_edges)
    if is_self.any():
        self_c, _ = np.histogram(rel_times[is_self], bins=bin_edges)
    else:
        self_c = np.zeros(len(bin_ctrs))
    hist_all[det_lid] = (counts, self_c)

fig, axes = plt.subplots(4, 4, figsize=(24, 18))
for det_lid in range(1, N_LASERS + 1):
    ax = axes[(det_lid - 1) // 4][(det_lid - 1) % 4]
    counts, self_c = hist_all[det_lid]
    ghost_c = counts - self_c
    ax.bar(dist_axis, self_c, width=dist_axis[1]-dist_axis[0],
           color='tab:red', label=u'\u81ea\u56de\u58f0', alpha=0.9)
    if ghost_c.max() > 0:
        ax.bar(dist_axis, ghost_c, width=dist_axis[1]-dist_axis[0],
               bottom=self_c, color='tab:blue', label=u'\u4e32\u6270(\u9b3c\u5f71)', alpha=0.6)
    ax.axvline(D_CUMUL, color='k', ls=':', lw=1.0, alpha=0.5)
    ax.set_title(f'Laser {det_lid}', fontsize=9)
    ax.set_ylim(0, max(counts.max()*1.4, 1))
    ax.set_xlim(0, 300)
    ax.tick_params(labelsize=6)
    ax.grid(alpha=0.15)

fig.text(0.5, 0.02, 'Distance [m] (1ns bin = 15cm; TOF 2000ns = 300m)', ha='center', fontsize=11)
fig.text(0.01, 0.5, 'Cumulative echo count', va='center', rotation='vertical', fontsize=11)
fig.legend(['Self(red)', 'Ghost(blue)'], loc='upper center', ncol=2, fontsize=10)
plt.suptitle(f'All 16 lasers cumulative TOF histogram (target D={D_CUMUL:.0f}m, 1ns bins)\nSelf(red) piles at {D_CUMUL:.0f}m; Ghost(blue) scattered = crosstalk', fontsize=14, y=0.96)
plt.tight_layout(rect=[0.02, 0.04, 0.98, 0.93])
plt.savefig('crosstalk_v03_cumulative_tof.png', dpi=110, bbox_inches='tight')
plt.show()

print('=' * 76)
print(f'All 16 lasers cumulative TOF (target D={D_CUMUL:.0f}m, 1ns bins):')
print(f'  Laser  Echoes  Self  Ghost  PeakDist[m]  GhostDists')
for det_lid in range(1, N_LASERS + 1):
    echoes = [r for r in res_cumul if r[3] == det_lid]
    n_self = sum(1 for r in echoes if r[0] == det_lid)
    counts, _ = hist_all[det_lid]
    peak_bin = counts.argmax()
    ghost_bins = np.where((counts > 0) & (abs(dist_axis - D_CUMUL) > 5))[0]
    gd = ','.join(f'{dist_axis[gb]:.0f}' for gb in ghost_bins[:5])
    print(f'  {det_lid:>5d} {len(echoes):>7d} {n_self:>5d} {len(echoes)-n_self:>6d} {dist_axis[peak_bin]:>12.1f} [{gd}]')
print()


CELL_SUMMARY = r'''# ============================================================================
# 总结
# ============================================================================
print("=" * 76)
print("crosstalk_sim_v02 总结")
print("=" * 76)
print()
print(f"1. 时序来源: {EXCEL_FILE} Group A (只看 A 组)")
print(f"   - {N_LASERS} 个激光器")
print(f"   - 共 {len(firing_events)} 次发光事件（{sum(1 for fe in firing_events if fe[1]=='A')} 次 A 组）")
print(f"   - 每激光器平均发光 {len(firing_events)/N_LASERS:.1f} 次")
print()
print("2. 编码参数")
print(f"   - KICK_SPACING = {KICK_SPACING*1e6:.1f} us（两个 tick 间隔）")
print(f"   - TOF 固定 = 2000ns (0.2us)")
print(f"   - tx_trig_dly: 1ns 步长（各激光器不同）")
print(f"   - tdelay: 1/12ns 步长（按 kick 定义）")
print()
print("3. 输出文件")
print("   - crosstalk_v02_echo_16ch.png: 16 通道回波接收图")
print("   - crosstalk_v03_timing_blocks.png: 时序色块图")
print("   - crosstalk_v03_cumulative_tof.png: 累积 TOF 波形")
print()
print("4. 使用说明")
print("   - 修改 D_SHOW 看不同距离的回波图")
print("   - 修改 demo_lasers 看不同激光器的累积 TOF")
print("   - 修改 LASER_ENCODING 调整编码参数（如 tx_trig_dly）")
print("   - 修改 KICK_SPACING 和 TOF 固定值")
'''


def code_cell(cid, src):
    return {"cell_type": "code", "id": cid, "metadata": {},
            "execution_count": None, "outputs": [],
            "source": src.splitlines(keepends=True)}

def md_cell(cid, src):
    return {"cell_type": "markdown", "id": cid, "metadata": {},
            "source": src.splitlines(keepends=True)}

cells = [
    md_cell("v02_header", CELL_HEADER),
    code_cell("imports", CELL_IMPORTS),
    code_cell("excel_parse", CELL_EXCEL_PARSE),
    code_cell("firing_times", CELL_FIRING_TIMES),
    code_cell("echo_scan", CELL_ECHO_SCAN),
    code_cell("plot_16ch_echo", CELL_16CH_ECHO),
    code_cell("plot_timing_blocks", CELL_TIMING_BLOCK),
    code_cell("plot_cumulative_tof", CELL_CUMULATIVE_TOF),
    code_cell("summary", CELL_SUMMARY),
]

nb = {
    "cells": cells,
    "metadata": {
        "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
        "language_info": {"name": "python", "version": "3"},
    },
    "nbformat": 4,
    "nbformat_minor": 5,
}

OUT = "crosstalk_sim_v03.ipynb"
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)
print(f"已生成 {OUT}，共 {len(cells)} 个 cell。")