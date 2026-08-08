# -*- coding: utf-8 -*-
"""
build_tcode_calculator_v2.py —— 生成 tcode_calculator_v2.ipynb

相对 v1：
  - 默认 v40 离散字母表模式（少档、100ns 预算）
  - 已知零残留表优先（v40 L5/L17 + v30 连续短码）
  - 搜索：字母表 min-conflicts 随机重启（同 search_24step_local.py）
  - 保留 continuous 模式（fast_search + 预算下压，同 v1）

缩写：
  XM（XtalkMark，串扰标记）
  TOF（Time of Flight，飞行时间）
  SEP（Separation，最小可分辨时间间隔）
"""
import json
import os

OUT_NB = "tcode_calculator_v2.ipynb"


def code_cell(cid, source):
    return {"cell_type": "code", "id": cid, "metadata": {},
            "execution_count": None, "outputs": [],
            "source": source.splitlines(keepends=True)}


def md_cell(cid, source):
    return {"cell_type": "markdown", "id": cid, "metadata": {},
            "source": source.splitlines(keepends=True)}


OVERVIEW = r"""# tcode 计算器 v2（离散字母表 / v40）

从 Excel 导入发光时序，搜索能**滤净**模组内鬼影/串扰的 `tx_trig_dly` 码表。

**验收标准（不可妥协）**：1~600m 扫描，XM 之后**鬼影残留 = 0 且真峰误杀 = 0**。

## 相对 v1 的变化

| 项目 | v1 | v2（本 notebook） |
|---|---|---|
| 编码 | 连续整数 `0..B` | 默认**离散字母表**（v40，少档抗展宽糊码） |
| 预算 | 压到 32ns（ratio=1.5） | 默认 **100ns** 总跨度，档数少 |
| 搜索 | fast_search 码差抽样 | **字母表 min-conflicts 随机重启** |
| 已知解 | v30 连续 32ns 表 | **v40 零残留表** + v30 连续表（continuous 模式） |

## 已确认零残留（gap=2, SEP=12, 本 Excel）

| XM.ratio | 模式 | 档数 / 字母表 | 预算 |
|---|---|---|---|
| **2.5** | 离散 L5 | `{0,25,50,75,100}` | 100ns |
| **1.5** | 离散 L17 | linspace 0..100 | 100ns |
| **1.5** | 连续 | — | **32ns**（v30，展宽风险大） |

`{0,24,48,72,96}` + ratio=1.5 可用 **系统枚举**验证（4 组独立、再笛卡尔积扫距）：
`python docs/tcode/exhaust_24step_r15.py` 或单组完整 kick 回溯 `enum_group_complete.py`。

| 步骤 | 内容 |
|---|---|
| 1 | 读 Excel，画 kick 栅格 |
| 2 | 全局参数字典：`SCENE` / `TIMING` / `XM` / `TCODE` / `SEARCH` |
| 3 | 阶段 A：已知零残留表；阶段 B：字母表随机搜或连续预算下压 |
| 4 | 码矩阵 / 栅格 / 残留诊断 |

缩写：XM（XtalkMark，串扰标记）、TOF（Time of Flight，飞行时间）、
SEP（Separation，峰可分辨最小间隔）。
"""


CELL_IMPORT = r'''# ============================================================================
# 0. 导入
# ============================================================================
import os, sys, time, copy
from collections import defaultdict
from itertools import combinations

import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import openpyxl
from matplotlib.patches import Rectangle

try:
    get_ipython().run_line_magic("matplotlib", "inline")
except Exception:
    try:
        matplotlib.use("module://matplotlib_inline.backend_inline")
    except Exception:
        pass

for _f in ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Source Han Sans SC"]:
    try:
        plt.rcParams["font.sans-serif"] = [_f]
        break
    except Exception:
        continue
plt.rcParams["axes.unicode_minus"] = False

def show_fig(fig=None):
    fig = fig or plt.gcf()
    try:
        from IPython.display import display
        display(fig)
    except Exception:
        pass
    plt.close(fig)


C_LIGHT = 2.99792458e8
NS = 1e-9
KICK_SPACING = 2.2e-6
TOF_WINDOW = 2000e-9
N_BINS = 2000

sys.path.insert(0, os.path.join("docs", "tcode"))
'''


CELL_EXCEL = r'''# ============================================================================
# 1. 从 Excel 导入发光时序
# ============================================================================
EXCEL_FILE = "Elephant 时序表.xlsx"
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb["长焦"]
COL_A = range(8, 24)

tdelay_A = {}
for c in COL_A:
    v = ws.cell(4, c).value
    tdelay_A[c - 8] = int(v) if v is not None else 0

laser_fires_raw = []
laser_ids = []
for r in range(7, 106):
    v = ws.cell(r, 6).value
    if v is None:
        continue
    try:
        lid = int(v)
    except (ValueError, TypeError):
        continue
    laser_ids.append(lid)
    for k, c in enumerate(COL_A):
        x = ws.cell(r, c).value
        if x is not None:
            laser_fires_raw.append((lid, k, int(x)))

laser_ids = sorted(set(laser_ids))
N_LASERS = len(laser_ids)
SHOT_KICKS = {lid: sorted(k for (l, k, tx) in laser_fires_raw if l == lid)
              for lid in laser_ids}
EXCEL_TX = {(l, k): tx for (l, k, tx) in laser_fires_raw}
N_ACC = int(np.median([len(v) for v in SHOT_KICKS.values()]))

def laser_color(laser_id):
    idx = (laser_id - 1) % max(N_LASERS, 1)
    return mcolors.hsv_to_rgb((idx / max(N_LASERS, 1), 0.65, 0.92))

print(f"1. 已读 {EXCEL_FILE} [长焦]：{N_LASERS} 激光器，"
      f"每激光器约 {N_ACC} 次 kick，共 {len(laser_fires_raw)} 次发光")
'''


CELL_GRID = r'''# ============================================================================
# 1b. 发光时序 kick 栅格（与 v30 图 V30-1 同风格）
# ============================================================================
def plot_kick_grid(label_of, title):
    fig, ax = plt.subplots(figsize=(12, 9))
    for lid in laser_ids:
        for k in SHOT_KICKS[lid]:
            ax.add_patch(plt.Rectangle(
                (k - 0.46, lid - 0.40), 0.92, 0.80,
                facecolor=laser_color(lid), edgecolor="k",
                linewidth=0.4, alpha=0.95, zorder=2))
            ax.text(k, lid, f"{label_of(lid, k)}", fontsize=6.5,
                    ha="center", va="center", color="k", zorder=3)
    ax.set_xticks(range(16)); ax.set_yticks(laser_ids)
    ax.set_xlim(-0.6, 15.6); ax.set_ylim(0.4, N_LASERS + 0.6)
    ax.set_xlabel("kick 序号")
    ax.set_ylabel("激光器编号")
    ax.set_title(title)
    ax.grid(alpha=0.25, zorder=0); ax.set_axisbelow(True)
    plt.tight_layout(); show_fig(fig)


def code_label(tx_ns, kick):
    return f"{tx_ns}+{tdelay_A.get(kick, 0)}/12"


plot_kick_grid(lambda lid, k: code_label(EXCEL_TX[(lid, k)], k),
               f"图 C2-1  Excel 发光时序栅格（{N_LASERS}×16）")
'''


CELL_PARAMS = r'''# ============================================================================
# 2. 全局参数字典（v40 离散字母表为主）
# ============================================================================
SCENE = {"excel_file": EXCEL_FILE, "sheet": "长焦"}

TIMING = {
    "use_tx_trig_dly": True,
    "use_delta_dly": False,
}

XM = {
    "ratio": 1.5,
    "use_pulse_width": False,
    "pulse_width_ns": 8.0,
    "require_zero_residual": True,
}

TCODE = {
    # "alphabet" = v40 离散档；"continuous" = v1 连续整数 0..B
    "mode": "alphabet",
    "consider_gap": True,
    "max_gap": 2,
    "sep_ns": 12,
    "budget_ns": 100,          # 字母表总跨度 [ns]
    "n_levels": None,          # None → 按 ratio 自动（2.5→5档，1.5→17档）
    "alphabet": None,          # None → linspace 或 ratio 默认；可设 [0,24,48,72,96]
}

SEARCH = {
    # ---- 字母表模式（min-conflicts 随机重启，同 search_24step_local.py）----
    "seconds": 300.0,          # 单轮搜索时限 [s]；加大可提高命中零残留概率
    "max_steps": 40000,
    "plateau": 2000,
    "seed": 20260730,
    "log_every": 100,
    # ratio=1.5 主字母表搜不到时，依次尝试备用字母表
    "fallback_alphabets": [
        [0, 25, 50, 75, 100],
        [0, 24, 48, 72, 96],
    ],
    # 系统枚举（非随机）：docs/tcode/exhaust_24step_r15.py
    "use_exhaust_script": False,
    # ---- continuous 模式（同 v1）----
    "budget_hint_ns": None,
    "budget_max_ns": 128,
    "budget_step_ns": 4,
    "seconds_per_budget": 120.0,
    "max_misses": 2,
    "short_steps": 5000,
    "short_plateau": 800,
    # ---- 验收 ----
    "eval_step_m": 5.0,
    "screen_step_m": 10.0,
    "diag_step_m": 10.0,
    "diag_max_show": 40,
}


def auto_n_levels():
    if TCODE["n_levels"] is not None:
        return int(TCODE["n_levels"])
    r = float(XM["ratio"])
    if r >= 2.5:
        return 5
    if r >= 1.5:
        return 17
    return 9


def make_alphabet():
    if TCODE["alphabet"] is not None:
        return np.asarray(TCODE["alphabet"], dtype=np.int32)
    budget = int(TCODE["budget_ns"])
    levels = auto_n_levels()
    if levels == 5 and float(XM["ratio"]) >= 2.5:
        return np.array([0, 25, 50, 75, 100], dtype=np.int32)
    return np.asarray(np.round(np.linspace(0, budget, levels)), dtype=np.int32)


def auto_budget_hint_continuous():
    if SEARCH["budget_hint_ns"] is not None:
        return int(SEARCH["budget_hint_ns"])
    r = float(XM["ratio"])
    if r >= 2.5:
        return 24
    if r >= 1.5:
        return 32
    return 80


ALPHABET = make_alphabet()
N_LEVELS = len(ALPHABET)

print("2. 全局参数")
print(f"  TCODE.mode={TCODE['mode']}  budget={TCODE['budget_ns']}ns  "
      f"档数={N_LEVELS}  字母表={list(map(int, ALPHABET))}")
print(f"  XM.ratio={XM['ratio']}  SEP={TCODE['sep_ns']}ns  gap={TCODE['max_gap']}")
print(f"  验收：1~600m 鬼影残留=0 且 真峰误杀=0")
'''


CELL_CORE = r'''# ============================================================================
# 3. 核心：约束 / 字母表搜索 / 连续搜索 / 验收
# ============================================================================
import solve_tcode as S
import gen_tcode_figures as G
import fast_search_v22 as F

G.LASER_IDS = list(laser_ids)
G.KICKS_OF = {l: list(SHOT_KICKS[l]) for l in laser_ids}
G.N_ACC = N_ACC
G.TX_EXCEL = dict(EXCEL_TX)
G.FIRES_RAW = list(laser_fires_raw)

S.LASER_IDS = list(laser_ids)
S.KICKS_OF = {l: list(SHOT_KICKS[l]) for l in laser_ids}
S.N_ACC = N_ACC
S.VARS = [(l, k) for l in laser_ids for k in SHOT_KICKS[l]]
S.VIDX = {v: i for i, v in enumerate(S.VARS)}

F.LASER_IDS = list(laser_ids)
F.KICKS_OF = {l: tuple(SHOT_KICKS[l]) for l in laser_ids}
F.VIDX = S.VIDX


def rebuild_constraints(gap, sep):
    S.CROSSTALK_MAX_GAP = gap
    S.set_sep(sep)
    S.CLASSES = S.build_classes(gap=gap)
    S.AVOID_PAIRS = S.build_avoid_true(gap=gap)
    S.VAR2CLS = [[] for _ in S.VARS]
    for ci, (_, prs) in enumerate(S.CLASSES):
        for ia, ib in prs:
            S.VAR2CLS[ia].append(ci)
            S.VAR2CLS[ib].append(ci)
    S.VAR2CLS = [sorted(set(c)) for c in S.VAR2CLS]
    S.VAR2AVOID = [[] for _ in S.VARS]
    for pi, (ia, ib) in enumerate(S.AVOID_PAIRS):
        S.VAR2AVOID[ia].append(pi)
        S.VAR2AVOID[ib].append(pi)
    return len(S.CLASSES), len(S.AVOID_PAIRS)


def load_table_file(path):
    ns = {}
    with open(path, encoding="utf-8") as f:
        exec(compile(f.read(), path, "exec"), ns)
    tbl = ns["TCODE_TABLE"]
    if any((l, k) not in tbl for l, k in S.VARS):
        return None
    u = np.array([tbl[(l, k)] for (l, k) in S.VARS], dtype=np.int32)
    meta = {
        "path": path,
        "alphabet": list(map(int, ns.get("TCODE_ALPHABET", sorted(set(map(int, u)))))),
        "n_levels": int(ns.get("TCODE_N_LEVELS", len(set(map(int, u))))),
        "budget": int(ns.get("TCODE_BUDGET_NS", int(u.max()))),
    }
    return u, meta


def known_table_candidates():
    """已核验或 v40 候选表，零残留优先。"""
    r = float(XM["ratio"])
    if TCODE["mode"] == "alphabet":
        if r >= 2.5:
            names = [
                "tcode_table_v40_r2.5_L5_100ns.py",
                "tcode_table_zero_r2.5_24ns.py",
            ]
        else:
            names = [
                "tcode_table_v40_r1.5_L17_100ns.py",
                "tcode_table_v40_r1.5_cont100ns.py",
                "tcode_table_zero_r1.5_32ns.py",
            ]
    else:
        if r >= 2.5:
            names = ["tcode_table_zero_r2.5_24ns.py",
                     "tcode_table_v22_r2.5_24ns.py"]
        else:
            names = ["tcode_table_zero_r1.5_32ns.py",
                     "tcode_table_zero_r1.5_36ns.py",
                     "tcode_table_v22_r1.5_56ns.py"]
    out = []
    for n in names:
        path = os.path.join("docs", "tcode", n)
        if not os.path.isfile(path):
            continue
        loaded = load_table_file(path)
        if loaded is None:
            continue
        out.append(loaded)
    return out


def evaluate_code(u, ratio, step_m, gap):
    fn = S.make_code_fn(u)
    fr = G.build_firings(fn)
    tot = dict(gb=0, ga=0, tb=0, kill=0, res_same=0, res_cross=0, res_mix=0)
    for D in np.arange(5.0, 601.0, step_m):
        r = G.simulate(D, fr, max_gap=gap, ratio=ratio)
        for k in tot:
            tot[k] += r[k]
    return tot


def is_zero(u, ratio, gap):
    ev = evaluate_code(u, ratio, SEARCH["screen_step_m"], gap)
    if ev["ga"] or ev["kill"]:
        return False, ev
    ev = evaluate_code(u, ratio, SEARCH["eval_step_m"], gap)
    return (ev["ga"] == 0 and ev["kill"] == 0), ev


def solve_alphabet_once(alphabet, rng, max_steps, plateau):
    alphabet = np.asarray(alphabet, dtype=np.int32)
    u = alphabet[rng.integers(0, len(alphabet), len(S.VARS))].astype(np.int32)
    cost = S.total_cost(u)
    best_cost, stale = cost, 0
    for _ in range(max_steps):
        if cost == 0:
            return u
        bad_c = [ci for ci in range(len(S.CLASSES)) if S.class_cost(u, ci) > 0]
        bad_a = [pi for pi in range(len(S.AVOID_PAIRS)) if S.avoid_cost_one(u, pi)]
        if not bad_c and not bad_a:
            return u
        if bad_a and (not bad_c or rng.random() < 0.45):
            pi = int(bad_a[rng.integers(0, len(bad_a))])
            ia, ib = S.AVOID_PAIRS[pi]
            vi = int(ia if rng.random() < 0.5 else ib)
        else:
            prs = S.CLASSES[int(rng.choice(bad_c))][1]
            vi = int(prs[int(rng.integers(0, len(prs)))][int(rng.integers(0, 2))])
        touched_c = S.VAR2CLS[vi]
        touched_a = S.VAR2AVOID[vi]
        base = (cost
                - sum(S.class_cost(u, ci) for ci in touched_c)
                - sum(S.avoid_cost_one(u, pi) for pi in touched_a))
        best, bestc = [], 1 << 30
        old = int(u[vi])
        for val in alphabet:
            av = 0
            for pi in touched_a:
                ia, ib = S.AVOID_PAIRS[pi]
                other = int(u[ib if ia == vi else ia])
                av += int(abs(int(val) - other) < S.SEP)
            c = (base
                 + sum(S._class_cost_with(u, ci, vi, int(val)) for ci in touched_c)
                 + av)
            if c < bestc:
                best, bestc = [int(val)], c
            elif c == bestc:
                best.append(int(val))
        pick = old if (bestc == cost and rng.random() < 0.2) else int(best[rng.integers(0, len(best))])
        u[vi] = pick
        cost = (base
                + sum(S.class_cost(u, ci) for ci in touched_c)
                + sum(S.avoid_cost_one(u, pi) for pi in touched_a))
        if cost < best_cost:
            best_cost, stale = cost, 0
        else:
            stale += 1
            if stale >= plateau:
                return None
    return None


def search_alphabet_zero(alphabet, ratio, gap, seconds, seed):
    """字母表 min-conflicts 随机重启 → 粗筛 → 细扫零残留。"""
    alphabet = np.asarray(alphabet, dtype=np.int32)
    t0 = time.time()
    n_feas = n_try = 0
    best_ga = None
    seen = set()
    seed_cur = int(seed)

    while time.time() - t0 < seconds:
        n_try += 1
        u = solve_alphabet_once(alphabet, np.random.default_rng(seed_cur),
                                SEARCH["max_steps"], SEARCH["plateau"])
        seed_cur += 1
        if u is None:
            continue
        key = tuple(int(x) for x in u)
        if key in seen:
            continue
        seen.add(key)
        n_feas += 1
        ok, ev = is_zero(u, ratio, gap)
        if ok:
            return u, {"feasible": n_feas, "tries": n_try, "eval": ev,
                       "secs": time.time() - t0, "best_ga": 0, "seed_end": seed_cur - 1}
        if best_ga is None or ev["ga"] < best_ga:
            best_ga = ev["ga"]
        if n_feas % SEARCH["log_every"] == 0:
            print(f"    … 字母表{list(map(int, alphabet))}  "
                  f"可行{n_feas}/{n_try} 唯一={len(seen)}  "
                  f"最好残留={best_ga}  seed={seed_cur-1}  "
                  f"[{time.time()-t0:.0f}s]", flush=True)
    return None, {"feasible": n_feas, "tries": n_try, "best_ga": best_ga,
                  "secs": time.time() - t0, "seed_end": seed_cur - 1}


def search_continuous_zero(budget, ratio, gap, sep, seconds, seed):
    """v1 策略：fast_search 码差抽样 + 短重启。"""
    groups = F.laser_groups()
    quads = F.valid_quads(budget, sep)
    rng = np.random.default_rng(seed)
    t0 = time.time()
    n_feas = n_try = 0
    best_ga = None
    short_steps = int(SEARCH["short_steps"])
    short_plateau = int(SEARCH["short_plateau"])

    while time.time() - t0 < seconds:
        n_try += 1
        u = None
        if quads:
            codes, ok = {}, True
            for ks, ls in groups:
                sol = None
                for _ in range(80):
                    sol = F.try_group(budget, sep, ks, ls, quads, rng)
                    if sol is not None:
                        break
                if sol is None:
                    ok = False
                    break
                codes.update(sol)
            if ok:
                u = np.zeros(len(S.VARS), dtype=np.int32)
                for (l, k), vi in S.VIDX.items():
                    u[vi] = codes[(l, k)]
                if S.total_cost(u) != 0:
                    u = None
        if u is None and (n_try % 20 == 0):
            u = S.solve(budget, rng, max_steps=short_steps, plateau=short_plateau)
        if u is None:
            continue
        n_feas += 1
        ok, ev = is_zero(u, ratio, gap)
        if ok:
            return u, {"feasible": n_feas, "tries": n_try, "eval": ev,
                       "secs": time.time() - t0, "best_ga": 0}
        if best_ga is None or ev["ga"] < best_ga:
            best_ga = ev["ga"]
    return None, {"feasible": n_feas, "tries": n_try, "best_ga": best_ga,
                  "secs": time.time() - t0}


def diagnose_residuals(u, ratio, gap, step_m, max_show):
    fn = S.make_code_fn(u)
    fr = G.build_firings(fn)
    rows = []
    for D in np.arange(5.0, 601.0, step_m):
        tof = 2.0 * D / C_LIGHT
        shot_idx = {l: {k: i for i, k in enumerate(SHOT_KICKS[l])} for l in laser_ids}
        H = {l: np.zeros((len(SHOT_KICKS[l]), N_BINS)) for l in laser_ids}
        truth = {l: np.zeros(N_BINS, dtype=bool) for l in laser_ids}
        src = {l: [[] for _ in range(N_BINS)] for l in laser_ids}
        for (ea, ek, et) in fr:
            t_echo = et + tof
            for (ra, rk, rt) in fr:
                if not (rt <= t_echo <= rt + TOF_WINDOW):
                    continue
                is_true = (ea == ra and ek == rk)
                if (not is_true) and abs(ea - ra) > gap:
                    continue
                b = int(np.clip(np.floor((t_echo - rt) / NS), 0, N_BINS - 1))
                H[ra][shot_idx[ra][rk], b] += 1.0
                if is_true:
                    truth[ra][b] = True
                else:
                    src[ra][b].append((ea, ek, rk, ea == ra))
        for l in laser_ids:
            add, mx = H[l].sum(axis=0), H[l].max(axis=0)
            for b in np.flatnonzero(add > 0.5):
                if truth[l][b]:
                    continue
                if mx[b] * ratio > add[b]:
                    continue
                kinds = src[l][b]
                same = any(ek == rk for _, ek, rk, _ in kinds)
                cross = any(ek != rk for _, ek, rk, _ in kinds)
                if same and cross:
                    kind = "混合(同kick+跨kick)"
                elif same:
                    kind = "同kick串扰"
                elif any(x for *_, x in kinds):
                    kind = "自身混叠"
                else:
                    kind = "跨kick串扰"
                rows.append({
                    "D_m": D, "recv": l, "bin": int(b),
                    "dist_m": b * NS * C_LIGHT / 2.0,
                    "ratio_eff": float(add[b] / max(mx[b], 1e-9)),
                    "kind": kind, "sources": kinds[:8],
                })
                if len(rows) >= max_show:
                    return rows
    return rows


def plot_tcode_matrix(table, vmax, title):
    fig, ax = plt.subplots(figsize=(12, 7))
    mat = np.full((N_LASERS, 16), np.nan)
    for i, lid in enumerate(laser_ids):
        for k in SHOT_KICKS[lid]:
            mat[i, k] = table[(lid, k)]
    im = ax.imshow(mat, origin="lower", aspect="auto", cmap="viridis",
                   extent=[-0.5, 15.5, 0.5, N_LASERS + 0.5],
                   vmin=0, vmax=max(vmax, 1))
    for i, lid in enumerate(laser_ids):
        for k in SHOT_KICKS[lid]:
            v = table[(lid, k)]
            ax.text(k, lid, str(v), ha="center", va="center", fontsize=7,
                    color="white" if v < vmax * 0.55 else "black")
    ax.set_xticks(range(16)); ax.set_yticks(laser_ids)
    ax.set_xlabel("kick"); ax.set_ylabel("激光器")
    ax.set_title(title)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="tx_trig_dly [ns]")
    plt.tight_layout(); show_fig(fig)


print("3. 核心函数就绪")
'''


CELL_SOLVE = r'''# ============================================================================
# 4. 求解：阶段 A 已知表 → 阶段 B 搜索
# ============================================================================
gap = TCODE["max_gap"] if TCODE["consider_gap"] else 15
sep = int(TCODE["sep_ns"])
ratio = float(XM["ratio"])

n_cls, n_av = rebuild_constraints(gap, sep)
F.GAP = gap

print("=" * 72)
print(f"模式={TCODE['mode']}  ratio={ratio}  SEP={sep}ns  gap={gap}")
print(f"字母表={list(map(int, ALPHABET))}  档数={N_LEVELS}  预算={TCODE['budget_ns']}ns")
print(f"散开类={n_cls}  避真={n_av}  变量={len(S.VARS)}")
print("=" * 72)

RESULT = {
    "ok": False, "u": None, "table": None, "budget": None,
    "source": None, "eval": None, "residuals": [], "trace": [],
    "alphabet_used": list(map(int, ALPHABET)),
    "n_levels": N_LEVELS,
}


def take(u, tag, ev, alphabet=None):
    RESULT.update(
        u=u.copy(),
        budget=int(u.max()),
        source=tag,
        eval=ev,
        ok=True,
        alphabet_used=alphabet or sorted(set(map(int, u))),
        n_levels=len(set(map(int, u))),
    )


# ---- 阶段 A：已知零残留表 ----
print("\n阶段 A · 已知表验收")
for u_k, meta in known_table_candidates():
    if S.total_cost(u_k) != 0:
        continue
    ok, ev = is_zero(u_k, ratio, gap)
    mark = "★零残留" if ok else f"残留 {ev['ga']}"
    print(f"  {os.path.basename(meta['path']):<40} "
          f"L={meta['n_levels']} max={meta['budget']:>3d}ns  {mark}")
    if ok:
        take(u_k, f"known:{os.path.basename(meta['path'])}", ev, meta["alphabet"])
        break

# ---- 阶段 B：搜索 ----
if not RESULT["ok"]:
    print("\n阶段 B · 搜索")
    if TCODE["mode"] == "alphabet":
        tried = []
        for alph in [ALPHABET] + [np.array(a, dtype=np.int32)
                                   for a in SEARCH.get("fallback_alphabets") or []
                                   if a is not None]:
            alph = np.asarray(alph, dtype=np.int32)
            key = tuple(map(int, alph))
            if key in tried:
                continue
            tried.append(key)
            print(f"\n  搜字母表 {list(map(int, alph))}  "
                  f"时限 {SEARCH['seconds']:.0f}s  seed={SEARCH['seed']} …")
            u, info = search_alphabet_zero(
                alph, ratio, gap, SEARCH["seconds"], SEARCH["seed"])
            if u is not None:
                print(f"  ★零残留  可行 {info['feasible']}/{info['tries']}  "
                      f"{info['secs']:.0f}s  seed→{info['seed_end']}")
                take(u, f"search:alph={list(map(int, alph))}", info["eval"],
                     list(map(int, alph)))
                break
            print(f"  未找到  可行 {info['feasible']}/{info['tries']}  "
                  f"最好残留 {info['best_ga']}  {info['secs']:.0f}s")
            RESULT["trace"].append((list(map(int, alph)), False, info))
        if not RESULT["ok"]:
            print("\n  提示：加大 SEARCH.seconds，或本地跑 "
                  "docs/tcode/search_24step_local.py --loop")
    else:
        hint = auto_budget_hint_continuous()
        lb = 2 * sep
        B = max(hint, lb)
        while B <= SEARCH["budget_max_ns"]:
            u, info = search_continuous_zero(
                B, ratio, gap, sep, SEARCH["seconds_per_budget"], SEARCH["seed"] + B)
            if u is not None:
                print(f"  B={B}ns ★零残留")
                take(u, f"search:B={B}", info["eval"])
                break
            print(f"  B={B}ns 失败  最好残留 {info['best_ga']}")
            B += SEARCH["budget_step_ns"]
        if RESULT["ok"]:
            misses = 0
            B = RESULT["budget"] - SEARCH["budget_step_ns"]
            while B >= lb and misses <= SEARCH["max_misses"]:
                u, info = search_continuous_zero(
                    B, ratio, gap, sep, SEARCH["seconds_per_budget"], SEARCH["seed"] + B)
                if u is not None:
                    take(u, f"compress:B={B}", info["eval"])
                    RESULT["trace"].append((B, True))
                    misses = 0
                else:
                    RESULT["trace"].append((B, False))
                    misses += 1
                B -= SEARCH["budget_step_ns"]

if RESULT["u"] is not None:
    RESULT["table"] = {(l, k): int(RESULT["u"][S.VIDX[(l, k)])]
                       for (l, k) in S.VARS}
    if not RESULT["ok"]:
        RESULT["residuals"] = diagnose_residuals(
            RESULT["u"], ratio, gap, SEARCH["diag_step_m"], SEARCH["diag_max_show"])
else:
    print("\n★ 未找到零残留码表。")

TCODE_TABLE = RESULT["table"]
TCODE_OK = RESULT["ok"]
print("\n结果:", "滤净成功" if TCODE_OK else "未滤净/无解",
      f"| {RESULT['source']} | max={RESULT['budget']}ns | "
      f"实际档数={RESULT['n_levels']}")
'''


CELL_REPORT = r'''# ============================================================================
# 5. 报告 + 码图
# ============================================================================
print("=" * 72)
print("tcode 计算器 v2 · 条件与结果")
print("=" * 72)

print(f"""
【模式】TCODE.mode = {TCODE['mode']}
  字母表 / 档数     : {RESULT['alphabet_used']}  (L={RESULT['n_levels']})
  预算跨度          : {TCODE['budget_ns']} ns
  来源              : {RESULT['source']}
  max(tx_trig_dly)  : {RESULT['budget']} ns

【XM】ratio={XM['ratio']}  验收步长={SEARCH['eval_step_m']}m
""")

if RESULT["eval"]:
    ev = RESULT["eval"]
    print(f"  鬼影 {ev['gb']} → 残留 {ev['ga']}  真峰误杀 {ev['kill']}")
    if TCODE_OK:
        print("  ★ 零残留验收通过")
    else:
        print("  ★ 未完全滤净")

if TCODE_TABLE:
    print("\n【码表】")
    for lid in laser_ids:
        parts = "  ".join(f"K{k}={TCODE_TABLE[(lid,k)]}" for k in SHOT_KICKS[lid])
        print(f"  L{lid:<2d}  {parts}")
    vmax = int(TCODE["budget_ns"]) if TCODE["mode"] == "alphabet" else (RESULT["budget"] or 100)
    plot_tcode_matrix(
        TCODE_TABLE, vmax,
        f"图 C2-2  tcode 码矩阵（L={RESULT['n_levels']}，"
        f"XM_RATIO={XM['ratio']}，滤{'净' if TCODE_OK else '未净'}）")
    plot_kick_grid(
        lambda lid, k: TCODE_TABLE[(lid, k)],
        f"图 C2-3  tcode 发光栅格（格内 = tx_trig_dly [ns]）")

if RESULT["residuals"]:
    print("\n残留鬼影（节选）：")
    for r in RESULT["residuals"][:10]:
        print(f"  D={r['D_m']:.0f}m recv=L{r['recv']} {r['kind']} ratio={r['ratio_eff']:.2f}")
elif TCODE_OK:
    print("\n无残留鬼影。")
'''


CELL_EXPORT = r'''# ============================================================================
# 6. （可选）导出
# ============================================================================
EXPORT = False
EXPORT_NAME = (f"tcode_table_calc_v2_r{XM['ratio']}_L{RESULT['n_levels']}_"
               f"{RESULT['budget']}ns.py")

if EXPORT and RESULT["u"] is not None and TCODE_OK:
    out = os.path.join("docs", "tcode", EXPORT_NAME)
    note = (f"tcode_calculator_v2; mode={TCODE['mode']} "
            f"alphabet={RESULT['alphabet_used']} source={RESULT['source']}")
    S.dump_table(RESULT["u"], out, TCODE["sep_ns"],
                 RESULT["budget"] or int(RESULT["u"].max()), note=note)
    with open(out, "a", encoding="utf-8") as f:
        f.write(f"\nTCODE_ALPHABET={RESULT['alphabet_used']}\n")
        f.write(f"TCODE_N_LEVELS={RESULT['n_levels']}\n")
    print(f"已导出 {out}")
else:
    print("未导出（需 EXPORT=True 且零残留）。")
'''


def main():
    cells = [
        md_cell("tc2_overview", OVERVIEW),
        code_cell("tc2_import", CELL_IMPORT),
        code_cell("tc2_excel", CELL_EXCEL),
        code_cell("tc2_grid", CELL_GRID),
        code_cell("tc2_params", CELL_PARAMS),
        code_cell("tc2_core", CELL_CORE),
        code_cell("tc2_solve", CELL_SOLVE),
        code_cell("tc2_report", CELL_REPORT),
        code_cell("tc2_export", CELL_EXPORT),
    ]
    nb = {
        "nbformat": 4,
        "nbformat_minor": 5,
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {"name": "python", "pygments_lexer": "ipython3"},
        },
        "cells": cells,
    }
    with open(OUT_NB, "w", encoding="utf-8") as f:
        json.dump(nb, f, ensure_ascii=False, indent=1)
    print(f"已生成 {OUT_NB}，共 {len(cells)} cell")


if __name__ == "__main__":
    main()
