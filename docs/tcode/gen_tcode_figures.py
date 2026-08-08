# -*- coding: utf-8 -*-
"""
gen_tcode_figures.py —— 生成 docs/tcode/tcode_scheme.md 里的全部插图
=====================================================================
运行方式（在仓库根目录）：
    python docs/tcode/gen_tcode_figures.py

图片输出到 docs/tcode/img/。所有图都是【为了讲清楚原理】的示意/仿真图：
    fig_01  串扰鬼影是怎么来的（时间轴 + 落点公式）
    fig_02  ★核心★ 码固定 vs 码逐 kick 变：XM 滤不掉 / 滤得掉
    fig_03  码矩阵 c[l][k]：Excel 现状 vs 线性同余码
    fig_04  某一对激光器的码差序列：全相同 vs 全不同
    fig_05  全部激光器对的"最大重复次数"热图（XM 能不能滤的判据）
    fig_06  码步长 step 必须 ≥ 峰宽，否则码不同也落同一 bin
    fig_07  跨 kick 混叠鬼影：多了 (k发−k收)·T_kick 项，线性同余码同样能散开
    fig_08  实际仿真：1~600m 距离扫描的鬼影残留率（Excel 码 vs 线性同余码）

缩写：
  XM（XtalkMark，串扰标记）      TOF（Time of Flight，飞行时间）
  IRF（Instrument Response Function，仪器响应函数）
  LCG（Linear Congruential，线性同余）
"""
import os
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch, FancyArrowPatch

for _f in ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Source Han Sans SC"]:
    try:
        matplotlib.rcParams["font.sans-serif"] = [_f]
        break
    except Exception:
        continue
matplotlib.rcParams["axes.unicode_minus"] = False

HERE = os.path.dirname(os.path.abspath(__file__))
IMG = os.path.join(HERE, "img")
os.makedirs(IMG, exist_ok=True)
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))

# ---------------------------------------------------------------------------
# 公共常数（与 crosstalk_sim_v13/v20 完全一致）
# ---------------------------------------------------------------------------
C_LIGHT      = 2.99792458e8
NS           = 1e-9
KICK_SPACING = 2.2e-6
TOF_WINDOW   = 2000e-9
D_UNAMBIG    = TOF_WINDOW * C_LIGHT / 2.0        # 300 m
N_BINS       = 2000
BIN_M        = NS * C_LIGHT / 2.0                # 1 bin = 1ns ≈ 0.15 m
XM_RATIO     = 1.6
P_PRIME      = 17                                # 线性同余码的素数模
CODE_STEP    = 8                                 # 码步长 [ns]，必须 ≥ 回波峰宽（见图6）
                                                 #   码最大值 = (P-1)×step = 128ns
                                                 #   ≤ kick 间隙 200ns，不会挤掉 TOF 窗

def save(fig, name):
    p = os.path.join(IMG, name)
    fig.savefig(p, dpi=130, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  已生成 {os.path.relpath(p, ROOT)}")


# ===========================================================================
# 读 Excel 时序（与 v13 相同）；读不到就用一个等价的合成时序，保证脚本能跑
# ===========================================================================
def load_timing():
    xlsx = os.path.join(ROOT, "Elephant 时序表.xlsx")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        ws = wb["长焦"]
        fires = []
        ids = []
        for r in range(7, 106):
            v = ws.cell(r, 6).value
            if v is None:
                continue
            try:
                lid = int(v)
            except (ValueError, TypeError):
                continue
            ids.append(lid)
            for k, c in enumerate(range(8, 24)):
                x = ws.cell(r, c).value
                if x is not None:
                    fires.append((lid, k, int(x)))
        ids = sorted(set(ids))
        if fires:
            print(f"  已读取 Excel 时序：{len(ids)} 激光器 / {len(fires)} 次发光")
            return ids, fires
    except Exception as e:
        print(f"  [警告] 读 Excel 失败（{e}），改用合成时序")
    # 合成兜底：16 激光器，每个在 4 个 kick 发光，tx 交替 0/50
    ids = list(range(1, 17))
    fires = [(l, (l - 1) % 4 + 4 * j, 50 * ((l - 1) % 2)) for l in ids for j in range(4)]
    return ids, fires


LASER_IDS, FIRES_RAW = load_timing()
N_LASERS = len(LASER_IDS)
KICKS_OF = {l: sorted(k for (a, k, t) in FIRES_RAW if a == l) for l in LASER_IDS}
N_ACC = int(np.median([len(v) for v in KICKS_OF.values()]))
TX_EXCEL = {(l, k): t for (l, k, t) in FIRES_RAW}


# ---------------------------------------------------------------------------
# 两套编码
# ---------------------------------------------------------------------------
def code_excel(lid, kick):
    """Excel 现状：每个激光器一个固定值（0 或 50 ns），不随 kick 变。"""
    return TX_EXCEL.get((lid, kick), 0)


def code_lcg(lid, kick, P=P_PRIME, step=CODE_STEP):
    """线性同余码： c[l][k] = ((l·(k+1)) mod P) · step   [ns]"""
    return ((lid * (kick + 1)) % P) * step


PG_QUAD   = 9      # 第二层（逐 kick 全局偏移）的模
GSTEP     = 8      # 第二层步长 [ns]


def code_lcg2(lid, kick, P=P_PRIME, step=CODE_STEP, Pg=PG_QUAD, gstep=GSTEP):
    """线性同余码 + 第二层【逐 kick 全局二次偏移】：
         c[l][k] = ((l·(k+1)) mod P)·step  +  ((k²) mod Pg)·gstep
       第二项对所有激光器【相同】：
         - 同 kick 串扰的码差里它完全抵消 → 不破坏第一层的完美性；
         - 自身混叠(a=b, k_发=k_收−1)的码差里它不抵消 → 把第一层治不了的那一类打散。
       为什么必须是【二次】：第一层沿 k 是线性的，同一激光器相邻 kick 的码差恒为
       −l (mod P)，是个常数 → 自身混叠根本没被散开。二次项的一阶差分随 k 变化，正好补上。"""
    return ((lid * (kick + 1)) % P) * step + ((kick * kick) % Pg) * gstep


CODES = {"excel": code_excel, "lcg": code_lcg, "lcg2": code_lcg2}
CODE_LABEL = {
    "excel": "Excel 现状（tx 固定 0/50ns）",
    "lcg": f"第一层 线性同余码 c=((l·(k+1)) mod {P_PRIME})×{CODE_STEP}ns",
    "lcg2": f"第一层+第二层 c=((l·(k+1)) mod {P_PRIME})×{CODE_STEP} + ((k²) mod {PG_QUAD})×{GSTEP} ns",
}


def gauss_pulse(n_bins, center, amp=1.0, fwhm=4.0):
    """一个回波脉冲（示意用的 IRF 形状，峰值 = amp）。"""
    x = np.arange(n_bins)
    s = fwhm / 2.355
    return amp * np.exp(-0.5 * ((x - center) / s) ** 2)


# ===========================================================================
# 图 1：串扰鬼影是怎么来的
# ===========================================================================
def fig_01():
    fig, axes = plt.subplots(2, 1, figsize=(14, 8.2))

    for row, (dtx, title) in enumerate([
            (0,  "情形①  两个激光器码相同（tx_A = tx_B = 0）"),
            (30, "情形②  发射方码比接收方大 30 ns（tx_A = 30, tx_B = 0）")]):
        ax = axes[row]
        tof = 1000.0                          # 物体 150m → 2D/c = 1000ns
        tB, tA = 0.0, float(dtx)              # 收(B)、发(A) 的发光时刻 [ns]

        # 接收窗
        ax.add_patch(plt.Rectangle((tB, 0.15), 2000, 0.7, facecolor="#dfe9f5",
                                   edgecolor="steelblue", lw=1.4, zorder=1))
        ax.text(tB + 20, 0.88, "B 的 TOF 接收窗（2000 ns）", fontsize=9,
                color="steelblue", va="bottom")

        # 发光竖线
        ax.plot([tB, tB], [0.15, 1.35], color="steelblue", lw=2.5, zorder=3)
        ax.text(tB, 1.38, f"B 发光\nt={tB:.0f}ns\n(也是 B 的测距零点)", fontsize=8.5,
                ha="center", va="bottom", color="steelblue")
        ax.plot([tA, tA], [0.15, 1.05], color="crimson", lw=2.5, zorder=3)
        ax.text(tA + (110 if dtx == 0 else 30), 1.07, f"A 发光\nt={tA:.0f}ns",
                fontsize=8.5, ha="center", va="bottom", color="crimson")

        # 回波
        eB, eA = tB + tof, tA + tof
        ax.plot([eB], [0.5], marker="v", ms=15, color="steelblue", zorder=5)
        ax.text(eB, 0.40, f"B 自己的回波\n到达 {eB:.0f}ns", fontsize=8.5, ha="center",
                va="top", color="steelblue")
        ax.plot([eA], [0.72], marker="v", ms=15, color="crimson", zorder=5)
        ax.text(eA, 0.80, f"A 的回波（对 B 就是串扰）\n到达 {eA:.0f}ns", fontsize=8.5,
                ha="center", va="bottom", color="crimson")

        # 测距结果
        ax.annotate("", xy=(eB, 0.26), xytext=(tB, 0.26),
                    arrowprops=dict(arrowstyle="<->", color="steelblue", lw=1.6))
        ax.text((tB + eB) / 2, 0.235, f"B 测到自己：rec_tof = {eB-tB:.0f}ns → "
                f"{(eB-tB)*NS*C_LIGHT/2:.1f} m  √ 正确",
                fontsize=9.5, ha="center", va="top", color="steelblue",
                bbox=dict(fc="white", ec="none", alpha=0.85, pad=1.5))
        ax.annotate("", xy=(eA, 1.0), xytext=(tB, 1.0),
                    arrowprops=dict(arrowstyle="<->", color="crimson", lw=1.6))
        ax.text((tB + eA) / 2, 1.03, f"B 把 A 的回波也算成自己的："
                f"rec_tof = {eA-tB:.0f}ns → {(eA-tB)*NS*C_LIGHT/2:.1f} m  × 鬼影",
                fontsize=9.5, ha="center", va="bottom", color="crimson",
                bbox=dict(fc="white", ec="none", alpha=0.85, pad=1.5))

        ax.set_xlim(-120, 2150); ax.set_ylim(0, 1.85)
        ax.set_yticks([]); ax.set_xlabel("时间 [ns]（以 B 的 kick 基线为 0）")
        ax.set_title(title + f"   →  鬼影偏移 {dtx} ns = "
                             f"{dtx*NS*C_LIGHT/2:.2f} m", fontsize=11)
        ax.grid(alpha=0.2, axis="x")

    fig.suptitle("图1  串扰鬼影是怎么来的：鬼影落点被【码差】平移，真回波不动\n"
                 "rec_tof(真) = 2D/c        rec_tof(串扰) = 2D/c + (tx_发 − tx_收)",
                 fontsize=13.5, y=1.02)
    plt.tight_layout()
    save(fig, "fig_01_ghost_origin.png")


# ===========================================================================
# 图 2 ★核心★：码固定 vs 码逐 kick 变
# ===========================================================================
def fig_02():
    n = 260                       # 只画 260 个 bin 的局部
    c_true = 130                  # 真目标峰位
    scen = [
        ("码【固定】：4 次的码差都是 +50ns", [50, 50, 50, 50], "#c0392b"),
        ("码【逐 kick 变】：4 次码差各不相同", [-48, 16, 60, -20], "#1e8449"),
    ]

    fig, axes = plt.subplots(3, 2, figsize=(16, 11),
                             gridspec_kw={"height_ratios": [2.2, 1.25, 1.15]})

    for col, (title, dtx_list, ccol) in enumerate(scen):
        shots = np.zeros((4, n))
        for s, d in enumerate(dtx_list):
            shots[s] += gauss_pulse(n, c_true)            # 真目标：每次都在同一位置
            shots[s] += gauss_pulse(n, c_true + d)        # 串扰：位置随码差走
        add = shots.sum(axis=0)
        mx = shots.max(axis=0)
        thr = mx * XM_RATIO
        xb = np.arange(n)

        # --- (上) 4 次 shot ---
        ax = axes[0][col]
        for s in range(4):
            ax.plot(xb, shots[s] + s * 1.25, lw=1.6,
                    color=plt.cm.viridis(s / 3), label=f"shot{s}")
            ax.axvline(c_true, color="0.8", lw=0.8, zorder=0)
            ax.annotate(f"串扰 Δ={dtx_list[s]:+d}ns", (c_true + dtx_list[s], s * 1.25 + 1.05),
                        fontsize=7.5, ha="center", color="crimson")
        ax.axvline(c_true, color="k", ls=":", lw=1.2)
        ax.text(c_true, 5.3, "真目标\n（4 次都在同一位置）", fontsize=9, ha="center",
                va="bottom", color="k")
        ax.set_ylim(-0.2, 6.1)
        ax.set_yticks([]); ax.set_ylabel("4 次 shot（纵向错开）")
        ax.set_title(f"{title}", fontsize=12, color=ccol, fontweight="bold")
        ax.legend(fontsize=8, ncol=4, loc="upper left"); ax.grid(alpha=0.2)

        # --- (中) hist_add vs 门限 ---
        ax = axes[1][col]
        ax.fill_between(xb, 0, add, color="steelblue", alpha=0.3)
        ax.plot(xb, add, color="steelblue", lw=2.0, label="hist_add（4 次累加）")
        ax.plot(xb, thr, color="purple", lw=2.0, ls="--",
                label=f"判据门限 = hist_max × {XM_RATIO}")
        ax.axhline(4, color="k", ls=":", lw=0.8)
        ax.text(2, 4.05, "真目标应有高度 = 4", fontsize=8, va="bottom")

        # 标出每个峰的 add / max / 判决
        peaks = [c_true] + [c_true + d for d in sorted(set(dtx_list))]
        for p in peaks:
            a, m = add[p], mx[p]
            drop = m * XM_RATIO > a
            ax.annotate(f"add={a:.1f}\nmax={m:.1f}\nadd/max={a/m:.1f}\n"
                        + ("→ 丢弃 √" if drop else "→ 保留"),
                        (p, a), fontsize=7.5, ha="center", va="bottom",
                        xytext=(0, 5), textcoords="offset points",
                        color=("crimson" if drop else "darkgreen"))
        ax.set_ylim(0, 6.6)
        ax.set_ylabel("计数")
        ax.legend(fontsize=8, loc="upper left"); ax.grid(alpha=0.2)

        # --- (下) XM 之后 ---
        ax = axes[2][col]
        after = add.copy()
        killed = 0
        for p in peaks:
            if mx[p] * XM_RATIO > add[p]:
                lo, hi = max(0, p - 12), min(n, p + 13)
                after[lo:hi] = 0.0
                killed += 1
        ax.fill_between(xb, 0, add, color="0.8", label="XM 前")
        ax.plot(xb, after, color=ccol, lw=2.4, label="XM 后")
        ax.plot([c_true], [after[c_true]], "*", ms=18, color="gold", mec="k", mew=0.8,
                zorder=6, label="真目标（存活）")
        ax.annotate(f"真目标保留，高度 {after[c_true]:.0f}", (c_true, after[c_true]),
                    fontsize=9, ha="center", va="bottom", xytext=(0, 12),
                    textcoords="offset points", color="darkgreen")
        for p in peaks:
            if p != c_true and mx[p] * XM_RATIO > add[p]:
                ax.plot([p], [add[p]], "x", ms=11, mew=2.2, color="k", zorder=6)
        ax.set_ylim(0, 5.6)
        ax.set_xlabel("bin（1 bin = 1 ns = 15 cm）")
        ax.set_ylabel("计数")
        n_ghost = len(set(dtx_list))
        ax.set_title(f"XM 之后：{n_ghost} 个鬼影峰里丢掉了 {killed} 个"
                     + ("   × 一个都没滤掉！" if killed == 0 else "   √ 全部滤掉")
                     + "（黑叉 = 被丢弃的鬼影峰）",
                     fontsize=11, color=ccol)
        ax.legend(fontsize=8, loc="upper left"); ax.grid(alpha=0.2)

    fig.suptitle("图2 ★核心★  同一条串扰，码固定就滤不掉、码逐 kick 变就滤得掉\n"
                 "左：串扰 4 次都落同一 bin → hist_add 累到 4 → add/max=4 → XM 认作真目标\n"
                 "右：串扰 4 次落 4 个不同 bin → 每个 bin 只累到 1 → add/max=1 → 被 XM 全部丢弃",
                 fontsize=13.5, y=1.005)
    plt.tight_layout(rect=[0, 0, 1, 0.965])
    save(fig, "fig_02_fixed_vs_coded.png")


# ===========================================================================
# 图 3：码矩阵 c[l][k]
# ===========================================================================
def fig_03():
    fig, axes = plt.subplots(1, 2, figsize=(17, 7))
    for ax, mode in zip(axes, ["excel", "lcg"]):
        fn = CODES[mode]
        M = np.full((N_LASERS, 16), np.nan)
        for i, l in enumerate(LASER_IDS):
            for k in KICKS_OF[l]:
                M[i, k] = fn(l, k)
        im = ax.imshow(M, cmap="viridis", aspect="auto", origin="lower",
                       extent=[-0.5, 15.5, LASER_IDS[0] - 0.5, LASER_IDS[-1] + 0.5])
        for i, l in enumerate(LASER_IDS):
            for k in KICKS_OF[l]:
                ax.text(k, l, f"{int(M[i,k])}", ha="center", va="center", fontsize=7,
                        color=("white" if M[i, k] < np.nanmax(M) * 0.55 else "black"))
        ax.set_xticks(range(16)); ax.set_yticks(LASER_IDS)
        ax.set_xlabel("kick 序号 k"); ax.set_ylabel("激光器编号 l")
        n_val = len(set(int(v) for v in M[~np.isnan(M)]))
        ax.set_title(f"{CODE_LABEL[mode]}\n格内 = tx_trig_dly [ns]，"
                     f"共用到 {n_val} 种码值", fontsize=11)
        fig.colorbar(im, ax=ax, fraction=0.04, label="tx_trig_dly [ns]")
    fig.suptitle("图3  码矩阵 c[l][k]：左边每一行都是一个常数（这就是 XM 失效的根源），"
                 "右边每一行沿 kick 不断变化", fontsize=13.5, y=1.02)
    plt.tight_layout()
    save(fig, "fig_03_code_matrix.png")


# ===========================================================================
# 图 4：某一对激光器的码差序列
# ===========================================================================
def pick_pair():
    """挑一对【相邻编号 且 共同发光 kick 最多】的激光器来演示。"""
    best, bestn = None, -1
    for a in LASER_IDS:
        for b in LASER_IDS:
            if a >= b or abs(a - b) > 2:
                continue
            common = sorted(set(KICKS_OF[a]) & set(KICKS_OF[b]))
            if len(common) > bestn:
                best, bestn = (a, b, common), len(common)
    return best


def fig_04():
    pair = pick_pair()
    if pair is None or len(pair[2]) == 0:
        print("  [跳过 fig_04] 没有共同发光的相邻激光器对")
        return
    a, b, common = pair
    fig, axes = plt.subplots(2, 1, figsize=(14, 7.5))
    for ax, mode in zip(axes, ["excel", "lcg"]):
        fn = CODES[mode]
        diffs = [fn(a, k) - fn(b, k) for k in common]
        uniq = len(set(diffs))
        seen = {}
        for i, (k, d) in enumerate(zip(common, diffs)):
            lvl = seen.get(d, 0)                       # 同一位置多次重合时逐层往上摞
            seen[d] = lvl + 1
            ax.plot([d], [0], "o", ms=16, color=plt.cm.viridis(i / max(len(common) - 1, 1)),
                    zorder=4)
            ax.annotate(f"kick {k}\nΔ={d:+d}ns\n({d*NS*C_LIGHT/2:+.2f} m)", (d, 0),
                        fontsize=8.5, ha="center", va="bottom",
                        xytext=(0, 16 + 34 * (lvl if uniq == 1 else i % 2)),
                        textcoords="offset points",
                        arrowprops=dict(arrowstyle="-", lw=0.7, color="0.5"))
        ax.axvline(0, color="k", lw=1.0, ls=":")
        ax.text(0, -0.055, "0 = 与真目标重合", fontsize=8, ha="center", va="top")
        ax.set_ylim(-0.12, 0.42); ax.set_yticks([])
        ax.set_xlim(-145, 145)
        ax.set_xlabel(f"码差 Δ = c[L{a}][k] − c[L{b}][k]  [ns]"
                      f"    （= L{a} 打到 L{b} 上的鬼影相对真目标的偏移）")
        verdict = ("× 4 次全重合 → 鬼影每次落同一 bin → hist_add 累到 4 → XM 滤不掉"
                   if uniq == 1 else
                   f"√ {uniq} 个值互不相同 → 鬼影散到 {uniq} 个 bin → 每个只累到 1 → XM 全滤掉")
        ax.set_title(f"{CODE_LABEL[mode]}   —— {len(common)} 个共同 kick，"
                     f"码差取到 {uniq} 个不同值\n{verdict}",
                     fontsize=11, color=("crimson" if uniq == 1 else "darkgreen"))
        ax.grid(alpha=0.25, axis="x")
    fig.suptitle(f"图4  以 L{a} → L{b} 这一对为例：码差序列决定鬼影落在几个 bin 上",
                 fontsize=13.5, y=1.0)
    plt.tight_layout()
    save(fig, "fig_04_code_diff.png")


# ===========================================================================
# 图 5：全部激光器对的最大重复次数热图
# ===========================================================================
def repeat_matrix(mode):
    """k_max[a][b] = 对 (a→b) 这条串扰，最多有几个 kick 给出【相同】的码差。
       k_max = 1 表示完美散开（XM 能全滤）；k_max = 4 表示完全重合（XM 全滤不掉）。"""
    fn = CODES[mode]
    M = np.full((N_LASERS, N_LASERS), np.nan)
    for i, a in enumerate(LASER_IDS):
        for j, b in enumerate(LASER_IDS):
            if a == b:
                continue
            common = sorted(set(KICKS_OF[a]) & set(KICKS_OF[b]))
            if not common:
                continue
            diffs = [fn(a, k) - fn(b, k) for k in common]
            vals, cnts = np.unique(diffs, return_counts=True)
            M[i, j] = cnts.max()
    return M


def fig_05():
    fig, axes = plt.subplots(1, 3, figsize=(19, 6.4),
                             gridspec_kw={"width_ratios": [1, 1, 0.9]})
    Ms = {}
    for ax, mode in zip(axes[:2], ["excel", "lcg"]):
        M = repeat_matrix(mode)
        Ms[mode] = M
        im = ax.imshow(M, cmap="RdYlGn_r", vmin=1, vmax=max(N_ACC, 2), aspect="auto",
                       origin="lower",
                       extent=[LASER_IDS[0] - 0.5, LASER_IDS[-1] + 0.5,
                               LASER_IDS[0] - 0.5, LASER_IDS[-1] + 0.5])
        for i, a in enumerate(LASER_IDS):
            for j, b in enumerate(LASER_IDS):
                if not np.isnan(M[i, j]):
                    ax.text(b, a, f"{int(M[i,j])}", ha="center", va="center", fontsize=6.5)
        # 画出"空间上不可忽略"的带（编号间隔 ≤ 2）
        for off in (-2.5, 2.5):
            ax.plot([LASER_IDS[0] - 0.5, LASER_IDS[-1] + 0.5],
                    [LASER_IDS[0] - 0.5 + off, LASER_IDS[-1] + 0.5 + off],
                    "b--", lw=1.4, alpha=0.8)
        ax.set_xlabel("接收激光器 b"); ax.set_ylabel("发射激光器 a")
        ax.set_xticks(LASER_IDS); ax.set_yticks(LASER_IDS)
        ax.tick_params(labelsize=7)
        band = [M[i, j] for i, a in enumerate(LASER_IDS) for j, b in enumerate(LASER_IDS)
                if a != b and abs(a - b) <= 2 and not np.isnan(M[i, j])]
        worst = int(max(band)) if band else 0
        ax.set_title(f"{CODE_LABEL[mode]}\n蓝虚线内 = 空间上不可忽略的对（编号间隔≤2）；"
                     f"该区最坏重复次数 = {worst}", fontsize=10.5)
        fig.colorbar(im, ax=ax, fraction=0.045, label="同一码差最多重复几个 kick")

    # 第三栏：柱状对比 + 判据说明
    ax = axes[2]
    labels = ["excel", "lcg"]
    xs = np.arange(1, N_ACC + 1)
    w = 0.38
    for t, mode in enumerate(labels):
        M = Ms[mode]
        band = [M[i, j] for i, a in enumerate(LASER_IDS) for j, b in enumerate(LASER_IDS)
                if a != b and abs(a - b) <= 2 and not np.isnan(M[i, j])]
        cnt = [sum(1 for v in band if int(v) == k) for k in xs]
        ax.bar(xs + (t - 0.5) * w, cnt, w, label=CODE_LABEL[mode].split("（")[0],
               color=("#c0392b" if mode == "excel" else "#1e8449"), edgecolor="k")
        for k, c in zip(xs, cnt):
            if c:
                ax.text(k + (t - 0.5) * w, c, str(c), ha="center", va="bottom", fontsize=8)
    ax.axvspan(0.5, XM_RATIO, color="green", alpha=0.12)
    ax.axvline(XM_RATIO, color="purple", lw=2, ls="--")
    ax.text(XM_RATIO, ax.get_ylim()[1] * 0.95, f" XM_RATIO={XM_RATIO}", color="purple",
            fontsize=10, va="top")
    ax.set_xticks(xs)
    ax.set_xlabel("该串扰对的最大重复次数 k")
    ax.set_ylabel("激光器对数（只统计编号间隔≤2 的）")
    ax.set_title("绿色区（k < XM_RATIO）= XM 能滤掉\n"
                 "紫线右侧 = XM 滤不掉，必须靠编码搬过去", fontsize=10.5)
    ax.legend(fontsize=9); ax.grid(alpha=0.25, axis="y")

    fig.suptitle("图5  能不能滤，一张图看完：每一对 (发,收) 的「最大重复次数」k\n"
                 "k < XM_RATIO 才滤得掉 —— 编码的唯一任务就是把所有格子压到 k=1",
                 fontsize=13.5, y=1.03)
    plt.tight_layout()
    save(fig, "fig_05_repeat_heatmap.png")


# ===========================================================================
# 图 6：码步长必须 ≥ 峰宽
# ===========================================================================
def fig_06():
    n = 120
    c0 = 40
    fwhm = 4.0                       # 回波峰宽 [bin] ≈ IRF 宽度
    cases = [(1, "step = 1 ns"), (3, "step = 3 ns"), (8, "step = 8 ns")]
    fig, axes = plt.subplots(2, 3, figsize=(17, 7.6),
                             gridspec_kw={"height_ratios": [1.5, 1]})
    for col, (step, name) in enumerate(cases):
        shots = np.zeros((4, n))
        for s in range(4):
            shots[s] = gauss_pulse(n, c0 + s * step, fwhm=fwhm)
        add, mx = shots.sum(axis=0), shots.max(axis=0)
        xb = np.arange(n)

        ax = axes[0][col]
        for s in range(4):
            ax.plot(xb, shots[s], lw=1.6, color=plt.cm.viridis(s / 3), label=f"shot{s}")
        ax.set_xlim(c0 - 20, c0 + 45); ax.set_ylim(0, 1.3)
        ax.set_title(f"{name}   （回波峰宽 FWHM = {fwhm:.0f} bin）", fontsize=11)
        ax.set_ylabel("单次 shot")
        ax.legend(fontsize=7.5, ncol=2); ax.grid(alpha=0.2)

        ax = axes[1][col]
        ax.fill_between(xb, 0, add, color="steelblue", alpha=0.3)
        ax.plot(xb, add, color="steelblue", lw=2.0, label="hist_add")
        ax.plot(xb, mx * XM_RATIO, color="purple", lw=2.0, ls="--",
                label=f"max × {XM_RATIO}")
        pk = int(np.argmax(add))
        ratio = add[pk] / mx[pk]
        drop = mx[pk] * XM_RATIO > add[pk]
        ax.set_xlim(c0 - 20, c0 + 45); ax.set_ylim(0, 4.4)
        ax.set_xlabel("bin")
        ax.set_ylabel("计数")
        ax.set_title(f"峰顶 add/max = {ratio:.2f}  →  "
                     + ("√ 判为串扰、滤掉" if drop else "× 滤不掉（4 次糊成一个峰）"),
                     fontsize=10.5, color=("darkgreen" if drop else "crimson"))
        ax.legend(fontsize=8); ax.grid(alpha=0.2)

    fig.suptitle("图6  码步长 step 必须 ≥ 回波峰宽：码值不同 ≠ 落到不同 bin\n"
                 "step 太小时 4 次的串扰糊成一个峰，hist_add 照样被累高，XM 依然滤不掉",
                 fontsize=13.5, y=1.02)
    plt.tight_layout()
    save(fig, "fig_06_step_vs_binwidth.png")


# ===========================================================================
# 图 7：跨 kick 混叠鬼影
# ===========================================================================
def fig_07():
    fig, axes = plt.subplots(2, 1, figsize=(14.5, 8.6))

    # --- 上：时间轴示意 ---
    ax = axes[0]
    Tk = 2200.0
    for k in range(3):
        ax.add_patch(plt.Rectangle((k * Tk, 0.3), 2000, 0.45, facecolor="#dfe9f5",
                                   edgecolor="steelblue", lw=1.2))
        ax.text(k * Tk + 1000, 0.77, f"kick {k} 的接收窗（2000ns）", fontsize=8.5,
                ha="center", va="bottom", color="steelblue")
        ax.plot([k * Tk, k * Tk], [0.3, 1.0], color="steelblue", lw=2)
    D = 480.0
    tof = 2 * D / C_LIGHT * 1e9              # 3202 ns
    ax.plot([0], [0.52], "^", ms=13, color="crimson")
    ax.annotate("", xy=(tof, 0.52), xytext=(0, 0.52),
                arrowprops=dict(arrowstyle="->", color="crimson", lw=2))
    ax.plot([tof], [0.52], "v", ms=15, color="crimson")
    ax.text(tof, 0.60, f"D={D:.0f}m 的回波，飞行 {tof:.0f}ns 后到达\n"
                       f"—— 已经越过 kick0 的窗，落进 kick1 的窗",
            fontsize=9, ha="center", va="bottom", color="crimson")
    ax.annotate("", xy=(tof, 0.36), xytext=(Tk, 0.36),
                arrowprops=dict(arrowstyle="<->", color="darkorange", lw=1.8))
    ax.text((tof + Tk) / 2, 0.30, f"被 kick1 算成 rec_tof = {tof-Tk:.0f}ns "
                                  f"→ 伪装成 {(tof-Tk)*NS*C_LIGHT/2:.0f} m",
            fontsize=9.5, ha="center", va="top", color="darkorange")
    ax.set_xlim(-200, 3 * Tk); ax.set_ylim(0.1, 1.35); ax.set_yticks([])
    ax.set_xlabel("时间 [ns]")
    ax.set_title("跨 kick 混叠：超过 300m 的回波落进【后面某个 kick】的窗里\n"
                 "rec_tof = 2D/c + (tx_发 − tx_收) − (k_收 − k_发)·T_kick", fontsize=11.5)
    ax.grid(alpha=0.2, axis="x")

    # --- 下：跨 kick 情形下两套码的码差散布 ---
    # --- 下：最难的一类 —— 自身混叠（同一个激光器自己上一个 kick 的光）---
    ax = axes[1]

    def self_diffs(fn, l):
        ks = KICKS_OF[l]
        return [fn(l, max(p)) - fn(l, kb) for kb in ks
                for p in [[k for k in ks if k < kb]] if p]

    # 挑第一层表现最差（自混叠码差重复最多）的那个激光器来演示
    b = min(LASER_IDS, key=lambda l: (len(set(self_diffs(code_lcg, l))), -l))
    for row, mode in enumerate(["lcg", "lcg2"]):
        fn = CODES[mode]
        vals = []
        for kb in KICKS_OF[b]:
            prev = [k for k in KICKS_OF[b] if k < kb]
            if prev:
                ka = max(prev)
                vals.append((kb, ka, fn(b, ka) - fn(b, kb)))
        uniq = len(set(v[2] for v in vals))
        y = 1.35 * (1 - row)
        seen = {}
        for i, (kb, ka, d) in enumerate(vals):
            lvl = seen.get(d, 0); seen[d] = lvl + 1
            ax.plot([d], [y], "o", ms=15, color=plt.cm.viridis(i / max(len(vals) - 1, 1)),
                    zorder=4)
            ax.annotate(f"收K{kb}←自己K{ka}\nΔ={d:+d}ns", (d, y), fontsize=8,
                        ha="center", va="bottom",
                        xytext=(0, 14 + 30 * (lvl if uniq < len(vals) else i % 2)),
                        textcoords="offset points",
                        arrowprops=dict(arrowstyle="-", lw=0.6, color="0.6"))
        good = (uniq == len(vals))
        ax.text(-395, y, ("只有第一层（线性）" if mode == "lcg" else "第一层 + 第二层（二次）")
                + f"\n{len(vals)} 次里取到 {uniq} 个不同值\n"
                + ("× 几乎全重合，滤不掉" if uniq <= 1
                   else ("√ 全散开" if good else "△ 部分散开")),
                fontsize=10, va="center", ha="left",
                color=("crimson" if uniq <= 1 else ("darkgreen" if good else "darkorange")))
    ax.axvline(0, color="k", ls=":", lw=1.0)
    ax.set_xlim(-400, 260); ax.set_ylim(-0.5, 3.1); ax.set_yticks([])
    ax.set_xlabel(f"自身混叠鬼影的码差 Δ = c[L{b}][上一个kick] − c[L{b}][本kick]  [ns]")
    ax.set_title(f"以 L{b} 自己上一个 kick 的光为例（自身混叠，是最难治的一类）\n"
                 f"L{b} 在连续 kick {KICKS_OF[b]} 上发光；第一层 c=l·(k+1) 沿 k 是【线性】的，"
                 "相邻 kick 的码差恒为 −l (mod P)，只能取到 2 个值 ⟹ 连续发光时必然重合；\n"
                 "加一个所有激光器共用的【二次】偏移 (k² mod Pg)·step，其一阶差分随 k 变化，才把它打散",
                 fontsize=10.5)
    ax.grid(alpha=0.25, axis="x")

    fig.suptitle("图7  第二类鬼影：跨 kick 距离混叠 —— 同一套编码顺带也能治",
                 fontsize=13.5, y=1.01)
    plt.tight_layout()
    save(fig, "fig_07_cross_kick.png")


# ===========================================================================
# 图 8：实际仿真 —— 距离扫描的鬼影残留率
# ===========================================================================
def build_firings(code_fn):
    fs = []
    for (lid, k, tx0) in FIRES_RAW:
        tx = code_fn(lid, k)
        t = k * KICK_SPACING + tx * NS
        fs.append((lid, k, t))
    return fs


def simulate(D, firings, max_gap=2, ratio=XM_RATIO):
    """返回 dict：鬼影峰数（XM 前/后）、真峰数、误杀数，以及残留鬼影峰的成因分类。
       模型与 crosstalk_sim_v20 一致：δ 回波、1ns bin、编号间隔>max_gap 的串扰忽略。"""
    tof = 2.0 * D / C_LIGHT
    shot_idx = {l: {k: i for i, k in enumerate(KICKS_OF[l])} for l in LASER_IDS}
    H = {l: np.zeros((len(KICKS_OF[l]), N_BINS)) for l in LASER_IDS}
    truth = {l: np.zeros(N_BINS, dtype=bool) for l in LASER_IDS}   # 该 bin 是否含真回波
    same_k = {l: np.zeros(N_BINS, dtype=bool) for l in LASER_IDS}  # 含同 kick 串扰
    cross_k = {l: np.zeros(N_BINS, dtype=bool) for l in LASER_IDS} # 含跨 kick 鬼影

    for (ea, ek, et) in firings:
        t_echo = et + tof
        for (ra, rk, rt) in firings:
            if not (rt <= t_echo <= rt + TOF_WINDOW):
                continue
            is_true = (ea == ra and ek == rk)
            if (not is_true) and abs(ea - ra) > max_gap:
                continue                                   # 空间可忽略的串扰不进探测器
            b = int(np.clip(np.floor((t_echo - rt) / NS), 0, N_BINS - 1))
            H[ra][shot_idx[ra][rk], b] += 1.0
            if is_true:
                truth[ra][b] = True
            elif ek == rk:
                same_k[ra][b] = True
            else:
                cross_k[ra][b] = True

    out = dict(gb=0, ga=0, tb=0, kill=0, res_same=0, res_cross=0, res_mix=0)
    for l in LASER_IDS:
        add, mx = H[l].sum(axis=0), H[l].max(axis=0)
        for b in np.flatnonzero(add > 0.5):
            drop = mx[b] * ratio > add[b]
            if truth[l][b]:
                out["tb"] += 1
                out["kill"] += int(drop)
            else:
                out["gb"] += 1
                if not drop:                               # 残留：查它是什么造成的
                    out["ga"] += 1
                    if same_k[l][b] and cross_k[l][b]:
                        out["res_mix"] += 1
                    elif same_k[l][b]:
                        out["res_same"] += 1
                    else:
                        out["res_cross"] += 1
    return out


def fig_08():
    Ds = np.arange(5.0, 601.0, 5.0)
    KEYS = ["gb", "ga", "tb", "kill", "res_same", "res_cross", "res_mix"]
    res = {}
    for mode in ["excel", "lcg", "lcg2"]:
        fr = build_firings(CODES[mode])
        acc = {k: [] for k in KEYS}
        for i, D in enumerate(Ds):
            r = simulate(D, fr)
            for k in KEYS:
                acc[k].append(r[k])
            if i % 40 == 0:
                print(f"    [{mode}] 扫描 {i+1}/{len(Ds)}  D={D:.0f}m ...")
        res[mode] = {k: np.asarray(v) for k, v in acc.items()}

    fig, axes = plt.subplots(3, 1, figsize=(15, 11.5), sharex=True)
    cols = {"excel": "#c0392b", "lcg": "#e67e22", "lcg2": "#1e8449"}
    SHORT = {"excel": "Excel 现状", "lcg": "第一层 线性同余",
             "lcg2": "第一层+第二层（推荐）"}

    ax = axes[0]
    for m in ["excel", "lcg", "lcg2"]:
        ax.plot(Ds, res[m]["gb"], "--", color=cols[m], lw=1.3, alpha=0.65,
                label=f"鬼影峰·XM 前（{SHORT[m]}）")
        ax.plot(Ds, res[m]["ga"], "-", color=cols[m], lw=2.4,
                label=f"鬼影峰·XM 后残留（{SHORT[m]}）")
    ax.axvline(D_UNAMBIG, color="r", ls=":", lw=1.3)
    ax.text(D_UNAMBIG, ax.get_ylim()[1] * 0.92, " 300m", color="r", fontsize=9)
    ax.set_ylabel("鬼影峰数（16 激光器合计）")
    ax.set_title("(a) 鬼影峰：XM 滤除前 vs 滤除后\n"
                 "注意两条虚线不重合：编码把原本挤在同一个 bin 的鬼影拆散了，"
                 "所以「峰数」反而变多——但每个峰都只累到 1，全都能被滤掉")
    ax.legend(fontsize=8.5, ncol=2); ax.grid(alpha=0.25)

    ax = axes[1]
    for m in ["excel", "lcg", "lcg2"]:
        r = res[m]["ga"] / np.maximum(res[m]["gb"], 1) * 100
        ax.plot(Ds, r, "-", color=cols[m], lw=2.0,
                label=f"{SHORT[m]}：平均残留率 {r.mean():.1f}%")
    ax.set_ylabel("鬼影残留率 [%]"); ax.set_ylim(-3, 105)
    ax.set_title("(b) 鬼影残留率 —— 这就是编码有没有用的唯一指标")
    ax.legend(fontsize=10); ax.grid(alpha=0.25)

    ax = axes[2]
    ax.plot(Ds, res["excel"]["tb"], "k--", lw=3.0, alpha=0.45, label="真目标峰·XM 前")
    for i, m in enumerate(["excel", "lcg", "lcg2"]):
        ax.plot(Ds, res[m]["tb"] - res[m]["kill"], "-", color=cols[m], lw=2.4 - i * 0.7,
                label=f"真目标峰·XM 后存活（{SHORT[m]}）")
        ax.plot(Ds, res[m]["kill"], ":", color=cols[m], lw=1.6, alpha=0.9,
                label=f"! 误杀（{SHORT[m]}）= 全程 0")
    ax.axvline(D_UNAMBIG, color="r", ls=":", lw=1.3)
    ax.set_xlabel("物体真实距离 D [m]"); ax.set_ylabel("真目标峰数")
    ax.set_title("(c) 真目标有没有被误杀（D>300m 时真回波本来就超窗丢失，与编码无关）")
    ax.legend(fontsize=8.5, ncol=2); ax.grid(alpha=0.25)

    fig.suptitle(f"图8  实测效果：1~600m 距离扫描（XM_RATIO={XM_RATIO}，"
                 f"δ 回波模型，编号间隔≤2 的串扰）", fontsize=13.5, y=0.995)
    plt.tight_layout(rect=[0, 0, 1, 0.975])
    save(fig, "fig_08_sweep_result.png")

    # 顺便把汇总数字打出来，写进 md 正文
    print("\n  【图8 汇总】")
    for m in ["excel", "lcg", "lcg2"]:
        gb, ga = res[m]["gb"].sum(), res[m]["ga"].sum()
        tb, kk = res[m]["tb"].sum(), res[m]["kill"].sum()
        print(f"    {m:>6}: 鬼影峰 {gb} → {ga}  (滤除率 {(gb-ga)/max(gb,1):.2%})   "
              f"真峰 {tb}，误杀 {kk} ({kk/max(tb,1):.2%})")
        rs, rc, rm = (res[m][k].sum() for k in ("res_same", "res_cross", "res_mix"))
        print(f"            残留成因： 纯同kick串扰 {rs}  纯跨kick混叠 {rc}  两者混合 {rm}")


# ===========================================================================
# 图 9：残留率 vs XM_RATIO —— 编码和阈值怎么配合
# ===========================================================================
def fig_09():
    Ds = np.arange(5.0, 601.0, 10.0)          # 步长放宽到 10m，因为要跑 3 码 × 4 阈值
    ratios = [1.6, 2.0, 2.5, 3.0]
    modes = ["excel", "lcg", "lcg2"]
    SHORT = {"excel": "Excel 现状", "lcg": "第一层 线性同余",
             "lcg2": "第一层+第二层（推荐）"}
    cols = {"excel": "#c0392b", "lcg": "#e67e22", "lcg2": "#1e8449"}

    data = {m: {"resid": [], "same": [], "cross": [], "kill": []} for m in modes}
    for m in modes:
        fr = build_firings(CODES[m])
        for ratio in ratios:
            t = dict(gb=0, ga=0, tb=0, kill=0, res_same=0, res_cross=0, res_mix=0)
            for D in Ds:
                r = simulate(D, fr, ratio=ratio)
                for k in t:
                    t[k] += r[k]
            data[m]["resid"].append(t["ga"] / max(t["gb"], 1) * 100)
            data[m]["same"].append(t["res_same"])
            data[m]["cross"].append(t["res_cross"] + t["res_mix"])
            data[m]["kill"].append(t["kill"] / max(t["tb"], 1) * 100)
        print(f"    [{m}] 残留率 = "
              + ", ".join(f"r={r}:{v:.2f}%" for r, v in zip(ratios, data[m]["resid"])))

    fig, axes = plt.subplots(1, 3, figsize=(18, 5.8))
    xx, w = np.arange(len(ratios)), 0.26

    ax = axes[0]
    for i, m in enumerate(modes):
        bars = ax.bar(xx + (i - 1) * w, data[m]["resid"], w, color=cols[m],
                      edgecolor="k", label=SHORT[m])
        for b, v in zip(bars, data[m]["resid"]):
            ax.text(b.get_x() + b.get_width() / 2, v, f"{v:.1f}", ha="center",
                    va="bottom", fontsize=7.5)
    ax.set_xticks(xx); ax.set_xticklabels([f"{r}" for r in ratios])
    ax.set_xlabel("XM_RATIO"); ax.set_ylabel("鬼影残留率 [%]")
    ax.set_title("(a) 编码 × 阈值：鬼影残留率\n"
                 "编码不动只调阈值，效果有限；先把编码做对才是正道", fontsize=10.5)
    ax.legend(fontsize=8.5); ax.grid(alpha=0.25, axis="y")

    ax = axes[1]
    for i, m in enumerate(modes):
        ax.bar(xx + (i - 1) * w, data[m]["same"], w, color=cols[m], edgecolor="k",
               label=f"{SHORT[m]}·同 kick 串扰残留")
        ax.bar(xx + (i - 1) * w, data[m]["cross"], w, bottom=data[m]["same"],
               color=cols[m], edgecolor="k", alpha=0.45, hatch="//",
               label=f"{SHORT[m]}·跨 kick/自身混叠残留")
    ax.set_xticks(xx); ax.set_xticklabels([f"{r}" for r in ratios])
    ax.set_xlabel("XM_RATIO"); ax.set_ylabel("残留鬼影峰数")
    ax.set_title("(b) 残留是哪一类造成的（实心=同 kick，斜纹=跨 kick/自身混叠）\n"
                 "第一层治同 kick，第二层治跨 kick 与自身混叠", fontsize=10.5)
    ax.legend(fontsize=6.8, ncol=1); ax.grid(alpha=0.25, axis="y")

    ax = axes[2]; ax.axis("off")
    txt = ["结论（δ 理想模型，1~600m 扫描）", "─" * 34, ""]
    for m in modes:
        txt.append(f"{SHORT[m]}：")
        for r, v, kk in zip(ratios, data[m]["resid"], data[m]["kill"]):
            txt.append(f"   ratio={r:<4} 残留 {v:6.2f}%    误杀 {kk:.2f}%")
        txt.append("")
    txt += ["注意：本模型里真目标每次必亮，所以",
            "     误杀恒为 0；真实光子受限系统中，",
            "     ratio 越大越容易杀掉弱真目标，",
            "     不能只靠抬 ratio 来凑指标。"]
    # 注意：不能用 family="monospace"，DejaVu Sans Mono 没有中文字形，会全变成方框
    ax.text(0.0, 1.0, "\n".join(txt), fontsize=9.5, va="top", linespacing=1.5)

    fig.suptitle("图9  编码与阈值的配合：先把编码做对，再谈阈值", fontsize=13.5, y=1.02)
    plt.tight_layout()
    save(fig, "fig_09_ratio_tradeoff.png")


# ===========================================================================
if __name__ == "__main__":
    print("生成 tcode 图解 ...")
    fig_01()
    fig_02()
    fig_03()
    fig_04()
    fig_05()
    fig_06()
    fig_07()
    fig_08()
    fig_09()
    print("全部完成。")
